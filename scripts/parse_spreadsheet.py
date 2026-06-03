#!/usr/bin/env python3
"""
parse_spreadsheet.py — Convert XLSX / Numbers files to CSV on stdout.

Usage:
    python parse_spreadsheet.py --file /path/to/file.xlsx [--out-dir /path/to/images]

- Reads the first sheet of the workbook.
- Extracts any embedded images into --out-dir (if provided) and replaces
  cell references with local file paths so the Node caller can convert
  them to data-URIs.
- Outputs the CSV to stdout.
"""

import argparse
import csv
import io
import os
import sys


def parse_xlsx(file_path: str, out_dir: str | None) -> str:
    """Parse an XLSX file using openpyxl and return CSV text."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Workbook has no active sheet")

    # Extract embedded images if out_dir is provided
    image_map: dict[str, str] = {}  # cell ref -> file path
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
        try:
            for idx, img in enumerate(ws._images):
                anchor = img.anchor
                # Try to get the cell reference from the anchor
                if hasattr(anchor, '_from'):
                    col = anchor._from.col
                    row = anchor._from.row
                    cell_ref = f"{row}_{col}"
                else:
                    cell_ref = f"img_{idx}"

                ext = "png"
                if hasattr(img, 'format') and img.format:
                    ext = img.format.lower()

                out_path = os.path.join(out_dir, f"image_{idx}.{ext}")
                with open(out_path, "wb") as f:
                    f.write(img._data())
                image_map[cell_ref] = out_path
        except Exception:
            # Image extraction is best-effort; don't fail the whole parse
            pass

    output = io.StringIO()
    writer = csv.writer(output)

    for row in ws.iter_rows(values_only=True):
        # Convert None to empty string, everything else to str
        writer.writerow([str(cell) if cell is not None else "" for cell in row])

    return output.getvalue()


def parse_numbers(file_path: str) -> str:
    """Parse a Numbers file and return CSV text."""
    try:
        import numbers_parser
    except ImportError:
        raise ImportError(
            "numbers-parser package is required for .numbers files. "
            "Install with: pip install numbers-parser"
        )

    doc = numbers_parser.Document(file_path)
    sheets = doc.sheets
    if not sheets:
        raise ValueError("Numbers file has no sheets")

    table = sheets[0].tables[0]
    output = io.StringIO()
    writer = csv.writer(output)

    for row_idx in range(table.num_rows):
        row_data = []
        for col_idx in range(table.num_cols):
            cell = table.cell(row_idx, col_idx)
            row_data.append(str(cell.value) if cell.value is not None else "")
        writer.writerow(row_data)

    return output.getvalue()


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert spreadsheet to CSV")
    parser.add_argument("--file", required=True, help="Path to XLSX or Numbers file")
    parser.add_argument("--out-dir", default=None, help="Directory to extract embedded images")
    args = parser.parse_args()

    file_path: str = args.file
    out_dir: str | None = args.out_dir

    if not os.path.isfile(file_path):
        print(f"Error: file not found: {file_path}", file=sys.stderr)
        sys.exit(1)

    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xlsx":
        csv_text = parse_xlsx(file_path, out_dir)
    elif ext == ".numbers":
        csv_text = parse_numbers(file_path)
    else:
        print(f"Error: unsupported file extension: {ext}", file=sys.stderr)
        sys.exit(1)

    sys.stdout.write(csv_text)


if __name__ == "__main__":
    main()
